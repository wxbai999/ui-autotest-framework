"""
数据加载器 — 统一从 JSON / YAML / Excel 文件中加载测试数据。

所有加载方法返回 list[dict]，每条 dict = 一个测试用例的参数集。
"""

from __future__ import annotations

import json
import os
from typing import Any, Dict, List

import yaml


class DataLoader:
    """多格式测试数据加载器"""

    @staticmethod
    def load(file_path: str) -> List[Dict[str, Any]]:
        """
        根据扩展名自动识别格式并加载。

        Args:
            file_path: 数据文件路径（支持 .json / .yaml / .yml / .xlsx）

        Returns:
            list[dict] — 每项为一个测试用例的参数
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Data file not found: {file_path}")

        ext = os.path.splitext(file_path)[1].lower()

        if ext == ".json":
            return DataLoader._load_json(file_path)
        elif ext in (".yaml", ".yml"):
            return DataLoader._load_yaml(file_path)
        elif ext == ".xlsx":
            return DataLoader._load_excel(file_path)
        else:
            raise ValueError(f"Unsupported data file format: {ext}")

    @staticmethod
    def _load_json(file_path: str) -> List[Dict[str, Any]]:
        with open(file_path, encoding="utf-8") as f:
            data = json.load(f)
        # 兼容 {"test_cases": [...]} 和直接 [...] 两种结构
        if isinstance(data, dict):
            for key in ("test_cases", "cases", "data"):
                if key in data:
                    return data[key]
        if isinstance(data, list):
            return data
        raise ValueError(f"JSON file must contain a list or 'test_cases' key: {file_path}")

    @staticmethod
    def _load_yaml(file_path: str) -> List[Dict[str, Any]]:
        with open(file_path, encoding="utf-8") as f:
            data = yaml.safe_load(f)
        if isinstance(data, dict):
            for key in ("test_cases", "cases", "data"):
                if key in data:
                    return data[key]
        if isinstance(data, list):
            return data
        raise ValueError(f"YAML file must contain a list or 'test_cases' key: {file_path}")

    @staticmethod
    def _load_excel(file_path: str) -> List[Dict[str, Any]]:
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError("openpyxl is required for Excel support. pip install openpyxl")

        wb = load_workbook(file_path, read_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # 第一行作为字段名
        headers = [str(h) for h in rows[0]]
        data: List[Dict[str, Any]] = []
        for row in rows[1:]:
            record = {}
            for i, value in enumerate(row):
                if i < len(headers):
                    record[headers[i]] = value
            data.append(record)

        wb.close()
        return data


# 便捷函数
def load_test_data(file_path: str) -> List[Dict[str, Any]]:
    return DataLoader.load(file_path)
